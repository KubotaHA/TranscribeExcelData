#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import traceback
import yaml
import re

try:
    import openpyxl as opyxl
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.styles.colors import Color
except:
    print("[ERROR] openpyxl がインストールされていない可能性があります。")
    print(traceback.format_exc())

class OpyxlWrapper:
    def __init__(self, exel_file) -> None:
        self.exel_file = exel_file
        self.workbook = None
        self.worksheet = None

    def load_workbook(self):
        try:
            self.workbook = opyxl.load_workbook(filename=self.exel_file)
        except Exception as e:
            raise Exception("[ERROR load_workbook] Unexpected Error has been ocurred. -> " + str(e))
        return self.workbook

    # read_only でworkbookを読み込むとセルの一部の属性が取得できなくなるため基本的には使用しないこととする
    # def load_workbook_readonly(self):
    #     try:
    #         self.workbook = opyxl.load_workbook(filename=self.exel_file, read_only=True)
    #     except Exception as e:
    #         raise Exception("[ERROR load_workbook_readonly] Unexpected Error has been ocurred. -> " + str(e))

    def load_worksheet(self, sheet_number=None, sheet_name=None)->object:
        if self.workbook == None:
            raise Exception("[ERROR: load_worksheet] Exel file has not been loaded yet.")
        try:
            # シート番号で指定
            if sheet_number != None and sheet_name == None:
                # シート番号の指定が0以下はエラーとする
                if sheet_number < 1:
                    raise Exception("[ERROR load_worksheet] Unexpected sheet number. [number > 0]")
                worksheet_title = self.workbook.worksheets[sheet_number-1].title
                self.worksheet = self.workbook[worksheet_title]
            # シート名で指定
            elif sheet_number == None and sheet_name != None:
                self.worksheet = self.workbook[sheet_name]
            # シート番号とシート名で指定 ※シート名を使用する
            elif sheet_number != None and sheet_name != None:
                self.worksheet = self.workbook[sheet_name]
            else:
                raise Exception("[ERROR: load_worksheet] sheet_number and sheet_name are None.")
        except Exception as e:
            raise Exception("[ERROR load_worksheet] Unexpected Error has been ocurred. -> " + str(e))

    def get_celldata(self, row_number, column_a):
        if self.worksheet == None:
            raise Exception("[ERROR: get_celldata] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            return self.worksheet[specified_cell].value
        except Exception as e:
            raise Exception("[ERROR get_celldata] Unexpected Error has been ocurred. -> " + str(e))

    def get_cellcolor(self, row_number, column_a)->str or int:
        if self.worksheet == None:
            raise Exception("[ERROR: get_cellcolor] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            # FIXME: cell.fillを返すようにしたい
            return (self.worksheet[specified_cell].fill.start_color.index,
                    self.worksheet[specified_cell].fill.fill_type)
        except Exception as e:
            raise Exception("[ERROR get_cellcolor] Unexpected Error has been ocurred. -> " + str(e))

    def get_cellalignment(self, row_number, column_a)->dict:
        if self.worksheet == None:
            raise Exception("[ERROR: get_cellalignment] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            return self.worksheet[specified_cell].alignment
        except Exception as e:
            raise Exception("[ERROR get_cellalignment] Unexpected Error has been ocurred. -> " + str(e))

    def _get_mergedcells(self)->object:
        if self.worksheet == None:
            raise Exception("[ERROR: _get_mergedcells] Exel file and worksheet have not been loaded yet.")
        try:
            return self.worksheet.merged_cells
        except Exception as e:
            raise Exception("[ERROR _get_mergedcells] Unexpected Error has been ocurred. -> " + str(e))

    def _is_mergedcell_correct_format(self, target_str):
        # "A3:AB4" のフォーマットであるかを正規表現で検索する
        re_result = re.search(r'^[A-Z]+[0-9]+:[A-Z]+[0-9]+$', target_str)
        if re_result == None:
            return False
        else:
            return True

    def get_mergedcells_list(self)->list:
        if self.worksheet == None:
            raise Exception("[ERROR: get_mergedcells_list] Exel file and worksheet have not been loaded yet.")
        try:
            merged_cells = self._get_mergedcells()
            merged_cells_list = []
            for merged_cell in merged_cells:
                merged_cell_char = str(merged_cell)
                if self._is_mergedcell_correct_format(merged_cell_char) == False:
                    raise Exception("merged_cell_char is incorrect. -> '{0}'".format(merged_cell_char))
                merged_cells_list.append(merged_cell_char.split(':'))
            return merged_cells_list
        except Exception as e:
            raise Exception("[ERROR get_mergedcells_list] Unexpected Error has been ocurred. -> " + str(e))

    def write_celldata(self, row_number, column_a, cell_value):
        if self.worksheet == None:
            raise Exception("[ERROR: write_celldata] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            self.worksheet[specified_cell].value = cell_value
        except Exception as e:
            raise Exception("[ERROR write_celldata] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.fills.html
    # https://openpyxl.readthedocs.io/en/stable/styles.html
    def fill_cellcolor(self, row_number, column_a, color_index, filltype="solid"):
        # FIXME: 正しく色が反映されないため何も処理しない
        try:
            # if self.worksheet == None:
            #     raise Exception("Exel file and worksheet have not been loaded yet.")
            # if type(color_index) == int:
            #     ## FIXME: 黒
            #     if color_index == 1:
            #         color_index = 0
            #     ## FIXME: 白
            #     elif color_index == 0:
            #         color_index = 1
            #     ## FIXME: 色なしにする
            #     else:
            #         color_index = '00000000'
            # # color_index: 0-63
            # if type(color_index) == int:
            #     my_color = Color(indexed=color_index)
            # # color_index: 00000000-FFFFFFFF
            # elif type(color_index) == str:
            #     my_color = Color(rgb=color_index)
            # else:
            #     raise Exception("[ERROR: fill_cellcolor] Unexpected color_index has been ocurred..")
            # specified_cell = column_a + str(row_number)
            # # Colorクラスでは'00000000'は「黒」を示すが、色なしセルをget_cellcolorした返り値が'00000000'であるため、
            # # ここでは「色なし」を filltype == None の場合のみ処理することとする。
            # # また、「黒」のセルについてget_cellcolorした場合の返り値は int型で 0 であるため、
            # # fill_cellcolorの color_index に int型で 0 を指定することでfill可能である
            # if color_index == "00000000" and filltype == None:
            #     self.worksheet[specified_cell].fill = PatternFill(fill_type=None)
            # else:
            #     self.worksheet[specified_cell].fill = PatternFill(patternType=filltype, fgColor=my_color)
            pass
        except Exception as e:
            raise Exception("[ERROR get_cellcollor] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html?highlight=Alignment#openpyxl.styles.alignment.Alignment
    # https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/#openpyxl-alignment
    # https://cercopes-z.com/Python/openpyxl/module-styles-alignment-opxl.html
    def set_alignment(self, row_number, column_a, horizontal, vertical, text_rotation, wrap_text, shrink_to_fit, indent, justifyLastLine, readingOrder):
        if self.worksheet == None:
            raise Exception("[ERROR: set_alignment] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            # 横位置を「標準」で定義する場合は "general" を指定するが、
            # get_cellalignmentした結果は 0 となるため、ここでは「標準」を 0 として処理することとする。
            if horizontal==None:
                horizontal="general"
            self.worksheet[specified_cell].alignment = Alignment(
                                                            horizontal=horizontal,
                                                            vertical=vertical,
                                                            text_rotation=text_rotation,
                                                            wrap_text=wrap_text,
                                                            shrink_to_fit=shrink_to_fit,
                                                            indent=indent,
                                                            justifyLastLine=justifyLastLine,
                                                            readingOrder=readingOrder
                                                            )
        except Exception as e:
            raise Exception("[ERROR set_alignment] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/usage.html#merge-unmerge-cells
    def merge_cells(self, target_cells):
        if self.worksheet == None:
            raise Exception("[ERROR: merge_cells] Exel file and worksheet have not been loaded yet.")
        try:
            self.worksheet.merge_cells(target_cells)
        except Exception as e:
            raise Exception("[ERROR merge_cells] Unexpected Error has been ocurred. target_cells='{target_cells}' ERROR -> {error}".format(
                                                                                                target_cells=target_cells, error=str(e)))

    def save_workbook(self):
        if self.workbook == None:
            raise Exception("[ERROR: save_workbook] Exel file has not been loaded yet.")
        try:
            self.workbook.save(self.exel_file)
        except PermissionError as pe:
            raise PermissionError("[ERROR save_workbook] ファイルへのアクセス権限がありません。-> " + str(pe))
        except Exception as e:
            raise Exception("[ERROR save_workbook] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html?highlight=close#openpyxl.workbook.workbook.Workbook.close
    def close_workbook(self):
        if self.workbook == None:
            raise Exception("[ERROR: close_workbook] Exel file has not been loaded yet.")
        try:
            self.workbook.close()
            del self.exel_file
            del self.worksheet
            del self.workbook
        except Exception as e:
            raise Exception("[ERROR close_workbook] Unexpected Error has been ocurred. -> " + str(e))

def check_if_file_exists(file_path)->bool:
    if os.path.exists(file_path):
        return True
    else:
        return False

def get_yaml_data(yaml_file)->dict:
    try:
        with open(yaml_file, encoding='utf8') as file:
            yaml_data = yaml.safe_load(file)
        return yaml_data
    except PermissionError as pe:
        raise PermissionError("[ERROR get_yaml_data] ファイルへのアクセス権限がありません。-> " + str(pe))
    except FileNotFoundError as fe:
        raise FileNotFoundError("[ERROR] get_yaml_data ファイルが存在しません。-> " + str(fe))
    except Exception as e:
            raise Exception("[ERROR get_yaml_data] Unexpected Error has been ocurred. -> " + str(e))

def readparent2writeparent(read_parent, read_row, write_row, write_column)->str:
    re_result = re.search(r'^([A-Z]+)([0-9]+)$', read_parent)
    if re_result == None:
        raise Exception("[ERROR readparent2writeparent] read_parent is incorrect format. -> " + str(read_parent))
    row_delta = write_row - read_row
    return write_column + str(int(re_result.group(2)) + row_delta)

def main():
    print("[INFO] 処理を開始します。")
    try:
        # fielddefinition.yml を読み込む
        fielddefinition = "fielddefinition.yml"
        if check_if_file_exists(fielddefinition) == False:
            raise Exception("[ERROR main] ファイル {file} が存在しません。".format(file=fielddefinition))
        yaml_data = get_yaml_data(fielddefinition)

        # 読み込み対象のExcelファイルを読み込む
        read_target_filename = yaml_data["read_target"]["file_name"]
        if check_if_file_exists(read_target_filename) == False:
            raise Exception("[ERROR main] ファイル {file} が存在しません。".format(file=read_target_filename))
        read_target_sheetnumber = yaml_data["read_target"].get("sheet_number", None)
        read_target_sheetname = yaml_data["read_target"].get("sheet_name", None)
        if read_target_sheetnumber == None and read_target_sheetname == None:
            raise Exception("[ERROR main] シート名かシート番号の指定が存在しません。")
        read_target_column = yaml_data["read_target"]["column_definition"]
        read_target_row = yaml_data["read_target"]["row_definition"]
        read_target_row_max = yaml_data["read_target"]["row_max"]
        opened_read_target = OpyxlWrapper(read_target_filename)
        opened_read_target.load_workbook()
        opened_read_target.load_worksheet(read_target_sheetnumber, read_target_sheetname)
        # 読み込み対象のExcelファイルから結合されたセルのリストを取得する
        mergedcells_list = opened_read_target.get_mergedcells_list()
        # 書き込み対象のExcelファイルを読み込む
        write_target_filename = yaml_data["write_target"]["file_name"]
        if check_if_file_exists(write_target_filename) == False:
            raise Exception("[ERROR main] ファイル {file} が存在しません。".format(file=write_target_filename))
        write_target_sheetnumber = yaml_data["write_target"].get("sheet_number", None)
        write_target_sheetname = yaml_data["write_target"].get("sheet_name", None)
        write_target_column = yaml_data["write_target"]["column_definition"]
        write_target_row = yaml_data["write_target"]["row_definition"]
        opened_write_target = OpyxlWrapper(write_target_filename)
        opened_write_target.load_workbook()
        opened_write_target.load_worksheet(write_target_sheetnumber, write_target_sheetname)
        # 転記先と転記元における対象カラムの数が合わない場合はエラー
        if len(read_target_column) != len(write_target_column):
            print("[ERROR main] 転記先と転記元における対象カラムの数が合いません。")

        # 列を軸としてループする
        ## 指定されたカラムの分だけループ処理する
        for column_num in range(len(read_target_column)):
            # 読み込み対象の行の開始番号を定義する
            read_row = read_target_row
            # 書き込み対象の行の開始番号を定義する
            write_row = write_target_row
            ## 行ループを開始する -> 処理する行の最大値までループする
            while read_row <= read_target_row_max:
                # 読み込み対象のセルを指定する
                read_column = read_target_column[column_num]
                # セルの値を取得する
                read_value = opened_read_target.get_celldata(read_row, read_column)
                # セルの色を取得する
                read_color, read_filltype = opened_read_target.get_cellcolor(read_row, read_column)
                # セルの書式設定を取得する
                read_alignment = opened_read_target.get_cellalignment(read_row, read_column)
                # 書き込み対象のカラムを指定する
                write_column = write_target_column[column_num]
                # "A1"の形式で読み込み対象のセルを定義する
                read_cell = read_column + str(read_row)
                # "A1"の形式で書き込み対象のセルを定義する
                write_cell = write_column + str(write_row)
                # セルのマージを実行する
                for merged_cells in mergedcells_list:
                    # merged_cellsリストの0番目がセル結合の開始セルとなるため
                    # それが読み込み対象のセルと一致している場合のみセル結合を行う
                    if merged_cells[0] == read_cell:
                        # merged_cellsリストの1番目がセル結合の終端セルとなるため
                        # その情報から書き込み対象セルの終端を求める
                        merge_cells_parent = write_cell + ":" + readparent2writeparent(merged_cells[1], read_target_row, write_target_row, write_column)
                        opened_write_target.merge_cells(merge_cells_parent)
                        print("[INFO] {needed_merge_cells} をセル結合しました。".format(needed_merge_cells=merge_cells_parent))
                        break
                # セルの値が空でない場合のみセルへの書き込みを実行する
                if read_value != None and read_value != "":
                    # セルの値を書き込む
                    opened_write_target.write_celldata(write_row, write_column, read_value)
                    # セルに色を適用する
                    opened_write_target.fill_cellcolor(write_row, write_column, read_color, read_filltype)
                    # セルの書式設定を適用する
                    opened_write_target.set_alignment(
                        write_row, write_column,
                        horizontal=read_alignment.horizontal,
                        vertical=read_alignment.vertical,
                        text_rotation=read_alignment.text_rotation,
                        wrap_text=read_alignment.wrap_text,
                        shrink_to_fit=read_alignment.shrink_to_fit,
                        indent=read_alignment.indent,
                        justifyLastLine=read_alignment.justifyLastLine,
                        readingOrder=read_alignment.readingOrder
                        )
                    print("read_color: " + str(read_color) + " type: " + str(type(read_color)))
                    print("[INFO] 書き込みファイルのセル: {col}{row} を処理しました。".format(col=write_column, row=write_row))
                # 行番号をインクリメントする
                read_row += 1
                write_row += 1

        # 保存して終了する
        opened_read_target.close_workbook()
        opened_write_target.save_workbook()
        opened_write_target.close_workbook()
        print("[INFO] すべての処理が正常に終了しました。")
        exitcode = 0

    except Exception as e:
        # 異常があった場合はセーブしない
        print("[ERROR] 処理が異常終了しました。")
        print(str(e))
        print(str(traceback.format_exc()))
        exitcode = 1
    
    finally:
        opened_write_target.close_workbook()
        sys.exit(exitcode)

if __name__ == "__main__":
    main()
