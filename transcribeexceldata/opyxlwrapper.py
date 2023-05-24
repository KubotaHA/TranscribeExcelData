#!/usr/bin/env python
# -*- coding: utf-8 -*-

import traceback
import re

try:
    import openpyxl as opyxl
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.styles.colors import Color
except:
    print("[ERROR] openpyxl がインストールされていない可能性があります。")
    print(traceback.format_exc())

class OpyxlWrapper:
    def __init__(self, exel_file) -> None:
        self.exel_file = exel_file
        self.workbook = None
        self.worksheet = None
        self.worksheet_title = None

    def load_workbook(self):
        try:
            self.workbook = opyxl.load_workbook(filename=self.exel_file)
            return self.workbook
        except opyxl.utils.exceptions.InvalidFileException as ife:
            raise Exception(
                "[ERROR load_workbook] サポートされていないファイル形式が指定されました。'{file}'はExcelファイルでない可能性があります。".format(
                    file=self.exel_file
                )
            )
        except Exception as e:
            raise Exception("[ERROR load_workbook] Unexpected Error has been ocurred. -> " + str(e))

    # read_only でworkbookを読み込むとセルの一部の属性が取得できなくなるため基本的には使用しないこととする
    # def load_workbook_readonly(self):
    #     try:
    #         self.workbook = opyxl.load_workbook(filename=self.exel_file, read_only=True)
    #     except Exception as e:
    #         raise Exception("[ERROR load_workbook_readonly] Unexpected Error has been ocurred. -> " + str(e))

    def load_worksheet(self, sheet_number=None, sheet_name=None)->object:
        if self.workbook == None:
            raise Exception("[ERROR: load_worksheet] Exel file has not been loaded yet.")
        try:
            # シート番号で指定
            if sheet_number != None and sheet_name == None:
                # シート番号の指定が0以下はエラーとする
                if sheet_number < 1:
                    raise Exception("[ERROR load_worksheet] Unexpected sheet number. [number > 0]")
                self.worksheet_title = self.workbook.worksheets[sheet_number-1].title
                self.worksheet = self.workbook[self.worksheet_title]
            # シート名で指定
            elif sheet_number == None and sheet_name != None:
                self.worksheet_title = sheet_name
                self.worksheet = self.workbook[sheet_name]
            # シート番号とシート名で指定 ※シート名を使用する
            elif sheet_number != None and sheet_name != None:
                self.worksheet_title = sheet_name
                self.worksheet = self.workbook[sheet_name]
            else:
                raise Exception("[ERROR: load_worksheet] sheet_number and sheet_name are None.")
        except IndexError as ie:
            raise Exception("[ERROR load_worksheet] 指定のシート番号 {number} が対象ファイル上に存在しません。 -> ".format(number=str(sheet_number)) + str(ie))
        except Exception as e:
            raise Exception("[ERROR load_worksheet] Unexpected Error has been ocurred. -> " + str(e))

    def get_celldata(self, row_number, column_a):
        if self.worksheet == None:
            raise Exception("[ERROR: get_celldata] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            return self.worksheet[specified_cell].value
        except Exception as e:
            raise Exception("[ERROR get_celldata] Unexpected Error has been ocurred. -> " + str(e))

    def get_cellcolor(self, row_number, column_a)->str or int:
        if self.worksheet == None:
            raise Exception("[ERROR: get_cellcolor] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            # FIXME: cell.fillを返すようにしたい
            return (self.worksheet[specified_cell].fill.start_color.index,
                    self.worksheet[specified_cell].fill.fill_type)
        except Exception as e:
            raise Exception("[ERROR get_cellcolor] Unexpected Error has been ocurred. -> " + str(e))

    def get_cellalignment(self, row_number, column_a)->dict:
        if self.worksheet == None:
            raise Exception("[ERROR: get_cellalignment] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            return self.worksheet[specified_cell].alignment
        except Exception as e:
            raise Exception("[ERROR get_cellalignment] Unexpected Error has been ocurred. -> " + str(e))

    def _get_mergedcells(self)->object:
        if self.worksheet == None:
            raise Exception("[ERROR: _get_mergedcells] Exel file and worksheet have not been loaded yet.")
        try:
            return self.worksheet.merged_cells
        except Exception as e:
            raise Exception("[ERROR _get_mergedcells] Unexpected Error has been ocurred. -> " + str(e))

    def _is_mergedcell_correct_format(self, target_str):
        # "A3:AB4" のフォーマットであるかを正規表現で検索する
        re_result = re.search(r'^[A-Z]+[0-9]+:[A-Z]+[0-9]+$', target_str)
        if re_result == None:
            return False
        else:
            return True

    def get_mergedcells_list(self)->list:
        if self.worksheet == None:
            raise Exception("[ERROR: get_mergedcells_list] Exel file and worksheet have not been loaded yet.")
        try:
            merged_cells = self._get_mergedcells()
            merged_cells_list = []
            for merged_cell in merged_cells:
                merged_cell_char = str(merged_cell)
                if self._is_mergedcell_correct_format(merged_cell_char) == False:
                    raise Exception("merged_cell_char is incorrect. -> '{0}'".format(merged_cell_char))
                merged_cells_list.append(merged_cell_char.split(':'))
            return merged_cells_list
        except Exception as e:
            raise Exception("[ERROR get_mergedcells_list] Unexpected Error has been ocurred. -> " + str(e))

    def write_celldata(self, row_number, column_a, cell_value):
        if self.worksheet == None:
            raise Exception("[ERROR: write_celldata] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            self.worksheet[specified_cell].value = cell_value
        except Exception as e:
            raise Exception("[ERROR write_celldata] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.fills.html
    # https://openpyxl.readthedocs.io/en/stable/styles.html
    def fill_cellcolor(self, row_number, column_a, color_index, filltype="solid"):
        # FIXME: 正しく色が反映されないため何も処理しない
        try:
            # if self.worksheet == None:
            #     raise Exception("Exel file and worksheet have not been loaded yet.")
            # if type(color_index) == int:
            #     ## FIXME: 黒
            #     if color_index == 1:
            #         color_index = 0
            #     ## FIXME: 白
            #     elif color_index == 0:
            #         color_index = 1
            #     ## FIXME: 色なしにする
            #     else:
            #         color_index = '00000000'
            # # color_index: 0-63
            # if type(color_index) == int:
            #     my_color = Color(indexed=color_index)
            # # color_index: 00000000-FFFFFFFF
            # elif type(color_index) == str:
            #     my_color = Color(rgb=color_index)
            # else:
            #     raise Exception("[ERROR: fill_cellcolor] Unexpected color_index has been ocurred..")
            # specified_cell = column_a + str(row_number)
            # # Colorクラスでは'00000000'は「黒」を示すが、色なしセルをget_cellcolorした返り値が'00000000'であるため、
            # # ここでは「色なし」を filltype == None の場合のみ処理することとする。
            # # また、「黒」のセルについてget_cellcolorした場合の返り値は int型で 0 であるため、
            # # fill_cellcolorの color_index に int型で 0 を指定することでfill可能である
            # if color_index == "00000000" and filltype == None:
            #     self.worksheet[specified_cell].fill = PatternFill(fill_type=None)
            # else:
            #     self.worksheet[specified_cell].fill = PatternFill(patternType=filltype, fgColor=my_color)
            pass
        except Exception as e:
            raise Exception("[ERROR get_cellcollor] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.alignment.html?highlight=Alignment#openpyxl.styles.alignment.Alignment
    # https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/#openpyxl-alignment
    # https://cercopes-z.com/Python/openpyxl/module-styles-alignment-opxl.html
    def set_alignment(self, row_number, column_a, horizontal, vertical, text_rotation, wrap_text, shrink_to_fit, indent, justifyLastLine, readingOrder):
        if self.worksheet == None:
            raise Exception("[ERROR: set_alignment] Exel file and worksheet have not been loaded yet.")
        try:
            specified_cell = column_a + str(row_number)
            # 横位置を「標準」で定義する場合は "general" を指定するが、
            # get_cellalignmentした結果は 0 となるため、ここでは「標準」を 0 として処理することとする。
            if horizontal==None:
                horizontal="general"
            self.worksheet[specified_cell].alignment = Alignment(
                                                            horizontal=horizontal,
                                                            vertical=vertical, 
                                                            text_rotation=text_rotation,
                                                            wrap_text=wrap_text,
                                                            shrink_to_fit=shrink_to_fit,
                                                            indent=indent,
                                                            justifyLastLine=justifyLastLine,
                                                            readingOrder=readingOrder
                                                            )
        except Exception as e:
            raise Exception("[ERROR set_alignment] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/usage.html#merge-unmerge-cells
    def merge_cells(self, target_cells):
        if self.worksheet == None:
            raise Exception("[ERROR: merge_cells] Exel file and worksheet have not been loaded yet.")
        try:
            self.worksheet.merge_cells(target_cells)
        except Exception as e:
            raise Exception("[ERROR merge_cells] Unexpected Error has been ocurred. target_cells='{target_cells}' ERROR -> {error}".format(
                         target_cells=target_cells, error=str(e)))

    def reset_workbook_properties(self):
        if self.workbook == None:
            raise Exception("[ERROR: reset_workbook_properties] Excel file has not been loaded yet.")
        try:
            # 作成者の削除：作成者=openpyxlの情報を削除
            self.workbook.properties.creator = None
            # 最終更新者の削除
            self.workbook.properties.lastModifiedBy = None
        except Exception as e:
            raise Exception("[ERROR reset_workbook_properties] Unexpected Error has been ocurred. -> " + str(e))

    def save_workbook(self):
        if self.workbook == None:
            raise Exception("[ERROR: save_workbook] Exel file has not been loaded yet.")
        try:
            self.workbook.save(self.exel_file)
        except PermissionError as pe:
            raise PermissionError("[ERROR save_workbook] ファイルへのアクセス権限がありません。-> " + str(pe))
        except Exception as e:
            raise Exception("[ERROR save_workbook] Unexpected Error has been ocurred. -> " + str(e))

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html?highlight=close#openpyxl.workbook.workbook.Workbook.close
    def close_workbook(self):
        if self.workbook == None:
            raise Exception("[ERROR: close_workbook] Exel file has not been loaded yet.")
        try:
            self.workbook.close()
            del self.exel_file
            del self.worksheet
            del self.workbook
        except Exception as e:
            raise Exception("[ERROR close_workbook] Unexpected Error has been ocurred. -> " + str(e))
